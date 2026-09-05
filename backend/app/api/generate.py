import csv
import io
import json

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook

from app.config import get_settings
from app.models import BatchGenerateRequest, BatchItem, GenerateRequest, GenerateResponse
from app.services.exporter import export_batch_zip, export_qr
from app.utils.validation import assert_max_size

router = APIRouter(prefix="/generate", tags=["generate"])


@router.post("", response_model=GenerateResponse)
async def generate_qr(request: GenerateRequest) -> GenerateResponse:
    return export_qr(request)


@router.post("/batch")
async def generate_batch(request: BatchGenerateRequest) -> Response:
    archive = export_batch_zip(request)
    return Response(
        archive,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="qr-batch.zip"'},
    )


@router.post("/batch/upload")
async def generate_batch_upload(file: UploadFile = File(...), options_json: str = "{}") -> Response:
    settings = get_settings()
    data = await file.read()
    assert_max_size(len(data), settings.max_upload_bytes)
    options = GenerateRequest.model_validate_json(options_json)
    items = _parse_batch_file(file.filename or "", data)
    request = BatchGenerateRequest(items=items, options=options)
    archive = export_batch_zip(request)
    return Response(
        archive,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="qr-batch.zip"'},
    )


def _parse_batch_file(filename: str, data: bytes) -> list[BatchItem]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
        return [BatchItem(name=row.get("name") or f"qr-{idx + 1}", value=row.get("url") or row.get("value") or "") for idx, row in enumerate(rows)]
    if lower.endswith(".json"):
        payload = json.loads(data)
        if not isinstance(payload, list):
            raise HTTPException(status_code=400, detail="Batch JSON must be an array")
        return [BatchItem.model_validate(item) for item in payload]
    if lower.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip().lower() for cell in rows[0]]
        items: list[BatchItem] = []
        for idx, row in enumerate(rows[1:]):
            record = dict(zip(headers, row))
            items.append(BatchItem(name=str(record.get("name") or f"qr-{idx + 1}"), value=str(record.get("url") or record.get("value") or "")))
        return items
    raise HTTPException(status_code=415, detail="Batch file must be CSV, JSON, or XLSX")
